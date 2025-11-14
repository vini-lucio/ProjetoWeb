from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill
from io import BytesIO


def arquivo_excel(conteudo: list[list] | list[dict], cabecalho: list = [], titulo: str = '',
                  nova_aba: openpyxl.Workbook | None = None, cabecalho_negrito: bool = False,
                  formatar_numero: tuple[list[str], int] | None = None,
                  formatar_data: list[str] | None = None,
                  ajustar_largura_colunas: bool = False):
    """Gera arquivo excel sem salvar.

    Parametros:
    -----------
    :conteudo [list[list] | list[dict]]: com o conteudo para planilha. Quando o conteudo for uma lista de dicionarios o cabeçalho sempre será as chaves do primeiro item do dicionario.
    :cabecalho [list, Default []]: com o cabecalho no caso do conteudo ser uma lista de listas.
    :titulo [str, Default '']: com o nome da aba.
    :nova_aba [Workbook | None, Default None]: acrescenta o conteudo em uma nova aba no Workbook informado.
    :cabecalho_negrito [bool, Default False]: boolano se o cabeçalho será criado em negrito.
    :formatar_numero [tuple[list[str], int], Default None]: uma tupla onde o primeiro elemento é uma lista com a letra das colunas que serão formatadas como numero (pontuação de milhar) e o segundo elemento é o numero de casas decimais.
    :formatar_data [list[str], Default None]: com uma lista com a letra das colunas que serão formatadas como data (DD/MM/YYYY).
    :ajustar_largura_colunas[bool, Default False]: booleano se a largura das colunas são ajustadas automaticamente (o ajuste nem sempre tem o tamanho ideal).

    Retorno:
    --------
    :Workbook: com o arquivo excel do conteudo."""
    if not nova_aba:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
    else:
        workbook = nova_aba
        worksheet = workbook.create_sheet()

    if worksheet:
        if titulo:
            worksheet.title = titulo

        if conteudo:
            if isinstance(conteudo[0], dict):
                cabecalho = list(conteudo[0].keys())
                conteudo = [[linha.get(chave) for chave in cabecalho] for linha in conteudo]  # type:ignore

        if cabecalho:
            worksheet.append(cabecalho)

        for linha in conteudo:
            worksheet.append(linha)

        if cabecalho_negrito:
            negrito = Font(bold=True)
            for celula in worksheet[1]:
                celula.font = negrito

        if formatar_numero:
            colunas, casas_decimais = formatar_numero
            formato = '#,##0.' + ''.join(['0' for _ in range(casas_decimais)])
            for coluna in colunas:
                for celula in worksheet[coluna]:
                    celula.number_format = formato

        if formatar_data:
            colunas = formatar_data
            formato = 'DD/MM/YYYY'
            for coluna in colunas:
                for celula in worksheet[coluna]:
                    celula.number_format = formato

        if ajustar_largura_colunas:
            for coluna in worksheet.columns:
                maior_largura = 0
                letra_coluna = coluna[0].column_letter  # type:ignore
                for celula in coluna:
                    if celula.value:
                        maior_largura = max(maior_largura, len(str(celula.value)))
                largura_ajustada = maior_largura + 2
                worksheet.column_dimensions[letra_coluna].width = largura_ajustada

    return workbook


def salvar_excel_temporario(workbook: openpyxl.Workbook) -> BytesIO:
    """Retorna o byte stream na posição inicial com o objeto salvo.

    Parametros:
    -----------
    :workbook [Workbook]: com o arquivo excel.

    Retorno:
    --------
    :BytesIO: com o arquivo salvo em memoria."""
    byte_stream = BytesIO()
    workbook.save(byte_stream)
    byte_stream.seek(0)
    return byte_stream


def gerar_cabecalho(campos):
    """Retorna uma lista de campos, onde os campos são campos de um queryset"""
    cabecalho = [campo for campo in campos]
    return cabecalho


def gerar_conteudo_excel(queryset, cabecalho):
    """Retorna uma lista com o conteudo das linhas de uma planilha, onde o cabeçalho é uma lista de campos do queryset"""
    conteudo = []
    for obj in queryset:
        linha = [getattr(obj, field) for field in cabecalho]  # type:ignore
        conteudo.append(linha)
    return conteudo


def somar_coluna_formatada(conteudo, titulo_aba, workbook, letra_coluna_soma, cabecalho_soma):
    """Soma a coluna informada e destaca o resultado com a cor verde.

    Parametros:
    -----------
    :conteudo [list[list]]: com o conteudo do arquivo (resultado da função gerar_conteudo_excel).
    :titulo_aba [str]: com o nome da aba que terá soma.
    :workbook [Workbook]: com o arquivo excel.
    :letra_coluna_soma [str]: com a letra da coluna que será somada (somente um coluna por vez).
    :cabecalho_soma [str]: com o cabelho para a celula com a soma."""
    total_linhas = len(conteudo)

    worksheet = workbook[titulo_aba]

    verde = PatternFill(start_color='00CC00', end_color='00CC00', fill_type='solid')
    negrito = Font(bold=True)

    celula = worksheet[f'{letra_coluna_soma}{total_linhas + 2}']
    celula.value = f'=SUM({letra_coluna_soma}2:{letra_coluna_soma}{total_linhas + 1})'
    celula.number_format = '#,##0.00'
    celula.fill = verde
    celula.font = negrito

    celula = worksheet[f'{letra_coluna_soma}{total_linhas + 3}']
    celula.value = cabecalho_soma
    celula.fill = verde
    celula.font = negrito


def arquivo_excel_response(arquivo_salvo_excel, nome_arquivo):
    """Gera o HTTP Response com o arquivo excel para download.

    Parametros:
    -----------
    :arquivo_salvo_excel [Workbook | BytesIO]: com o arquivo excel.
    :nome_aquivo [str]: com o nome do arquivo para download.

    Retorno:
    --------
    :HttpResponse: com o arquivo para download."""
    response = HttpResponse(
        arquivo_salvo_excel,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="{}.xlsx"'.format(nome_arquivo)
    return response
